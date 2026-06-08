#!/usr/bin/env python3
"""Convert PRE-worldcup2.xlsx (batch 2) into a clean 7-column dataset CSV.

The xlsx carries the right schema (category_id, country, question, answer, options,
difficulty, explanation) but the data has two corruptions, both of which we repair
here so the output matches Pre-worldcup-clean.csv and can feed qa_review.py:

  1. `options` uses the same broken quoting batch 1 had — `[opt1",  "opt2",  "opt3"]"`.
     We reuse repair_worldcup.extract_options (the proven recovery for that pattern).
  2. Text fields are UTF-8 bytes that were decoded as Latin-1 (mojibake: "Mané" -> "ManÃ©").
     We undo it with a latin-1 -> utf-8 round-trip, guarded so clean/ASCII text is untouched.
"""
import openpyxl, csv, json
from repair_worldcup import extract_options

SRC = "PRE-worldcup2.xlsx"
OUT = "Pre-worldcup2.csv"
TEXT_COLS = (1, 2, 3, 5, 6)  # country, question, answer, difficulty, explanation


def demojibake(s):
    """Undo a UTF-8-as-Latin-1 misdecode; return s unchanged if it doesn't apply."""
    if not isinstance(s, str):
        return "" if s is None else str(s)
    try:
        fixed = s.encode("latin-1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return s
    # only accept the repair if it removed a mojibake marker
    return fixed if ("Ã" in s or "Â" in s) and "Ã" not in fixed else s


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    next(it)  # header
    n = bad_opts = 0
    with open(OUT, "w", newline="", encoding="utf-8") as g:
        w = csv.writer(g)
        w.writerow(["category_id", "country", "question", "answer",
                    "options", "difficulty", "explanation"])
        for r in it:
            r = list(r[:7]) + [""] * max(0, 7 - len(r))
            cat = str(r[0]) if r[0] is not None else "12"
            country, question, answer, diff, expl = (demojibake(r[i]) for i in TEXT_COLS)
            opts_blob = demojibake(r[4])
            options = extract_options(opts_blob)
            if len(options) != 3:
                bad_opts += 1
            w.writerow([cat, country, question, answer,
                        json.dumps(options, ensure_ascii=False), diff, expl])
            n += 1
    print("wrote %s: %d rows" % (OUT, n))
    print("rows whose options didn't recover to exactly 3 (TC-05 will flag): %d" % bad_opts)


if __name__ == "__main__":
    main()
